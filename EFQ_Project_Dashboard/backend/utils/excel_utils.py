from __future__ import annotations

from datetime import date, datetime, timezone
from pathlib import Path
from shutil import copy2
from tempfile import NamedTemporaryFile
from threading import RLock
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from utils.security import hash_password, now_iso


SHEET_SCHEMAS: dict[str, list[str]] = {
    'Users': [
        'UserID', 'FullName', 'NTID', 'Phone', 'Email', 'Role', 'PasswordHash',
        'AccountStatus', 'CreatedAt', 'UpdatedAt', 'LastLoginAt',
    ],
    'Incidents': [
        'IncidentID', 'IncidentTitle', 'Date', 'OEM', 'CustomerComplaint', 'DealerName',
        'DealerLocation', 'DealerContact', 'VehicleModel', 'VehicleVariant',
        'VehicleApplication', 'VIN', 'KilometerReading', 'ECUPartNumber', 'ECUName',
        'Severity', 'IssueType', 'CustodianNTID', 'CustodianName', 'Description',
        'Status', 'CreatedBy', 'CreatedAt', 'UpdatedBy', 'UpdatedAt',
    ],
    'Resolutions': [
        'ResolutionID', 'IncidentID', 'InvestigationDetails', 'RootCause', 'Recommendation',
        'ProposedSolution', 'CorrectiveAction', 'PreventiveAction', 'ValidationMethod',
        'ValidationResult', 'ValidationDate', 'ResolutionOwner', 'TargetDate',
        'ResolutionDate', 'ResolutionStatus', 'Remarks', 'CreatedAt', 'UpdatedAt',
    ],
    'Activities': ['ActivityID', 'IncidentID', 'Action', 'UserNTID', 'UserName', 'Timestamp'],
    'OEMRegions': ['OEM', 'Region', 'IsActive'],
    'ComplaintSuggestions': ['SuggestionID', 'ComplaintDescription', 'Category', 'IsActive'],
    'ECUs': ['ECUID', 'ECUName', 'ECUPartNumber', 'IsActive'],
    'DetectionPhases': ['PhaseID', 'PhaseName', 'IsActive'],
}


def _sample_users() -> list[dict[str, Any]]:
    timestamp = now_iso()
    default_password = hash_password('Passw0rd!')
    return [
        {'UserID': 'USR-000001', 'FullName': 'Avery Stone', 'NTID': 'ADM001', 'Phone': '+1555100001', 'Email': 'avery.stone@example.com', 'Role': 'Admin', 'PasswordHash': default_password, 'AccountStatus': 'Active', 'CreatedAt': timestamp, 'UpdatedAt': timestamp, 'LastLoginAt': ''},
        {'UserID': 'USR-000002', 'FullName': 'Jordan Vale', 'NTID': 'MGR001', 'Phone': '+1555100002', 'Email': 'jordan.vale@example.com', 'Role': 'Manager', 'PasswordHash': default_password, 'AccountStatus': 'Active', 'CreatedAt': timestamp, 'UpdatedAt': timestamp, 'LastLoginAt': ''},
        {'UserID': 'USR-000003', 'FullName': 'Engineer Ada', 'NTID': 'CUS001', 'Phone': '+1555100003', 'Email': 'engineer.ada@example.com', 'Role': 'Custodian', 'PasswordHash': default_password, 'AccountStatus': 'Active', 'CreatedAt': timestamp, 'UpdatedAt': timestamp, 'LastLoginAt': ''},
        {'UserID': 'USR-000004', 'FullName': 'Engineer Kai', 'NTID': 'CUS002', 'Phone': '+1555100004', 'Email': 'engineer.kai@example.com', 'Role': 'Custodian', 'PasswordHash': default_password, 'AccountStatus': 'Active', 'CreatedAt': timestamp, 'UpdatedAt': timestamp, 'LastLoginAt': ''},
        {'UserID': 'USR-000005', 'FullName': 'Taylor Reed', 'NTID': 'OTH001', 'Phone': '+1555100005', 'Email': 'taylor.reed@example.com', 'Role': 'Other', 'PasswordHash': default_password, 'AccountStatus': 'Active', 'CreatedAt': timestamp, 'UpdatedAt': timestamp, 'LastLoginAt': ''},
    ]


def _sample_incidents() -> list[dict[str, Any]]:
    return [
        {'IncidentID': 'EFQ-20260824-0001', 'IncidentTitle': 'Harness terminal intermittency on startup', 'Date': '2026-08-24', 'OEM': 'Atlas Motors', 'CustomerComplaint': 'Vehicle fails to crank consistently during first ignition attempt.', 'DealerName': 'North Ridge Auto Hub', 'DealerLocation': 'Pune', 'DealerContact': '+919800000001', 'VehicleModel': 'Aster X', 'VehicleVariant': 'Diesel AT', 'VehicleApplication': 'SUV', 'VIN': 'MA1ASTERX00000001', 'KilometerReading': 218, 'ECUPartNumber': 'ECU-AX-1001', 'ECUName': 'Powertrain Control Module', 'Severity': 'Critical', 'IssueType': '0 KM', 'CustodianNTID': 'CUS001', 'CustodianName': 'Engineer Ada', 'Description': 'Dealer reported intermittent no-crank behavior after overnight parking.', 'Status': 'Validation', 'CreatedBy': 'ADM001', 'CreatedAt': '2026-08-24T08:30:00+00:00', 'UpdatedBy': 'CUS001', 'UpdatedAt': '2026-08-25T11:15:00+00:00'},
        {'IncidentID': 'EFQ-20260825-0001', 'IncidentTitle': 'Cluster reset while driving over rough surface', 'Date': '2026-08-25', 'OEM': 'Nova Mobility', 'CustomerComplaint': 'Instrument cluster resets and warning lamps flash on rough roads.', 'DealerName': 'Skyline Wheels', 'DealerLocation': 'Chennai', 'DealerContact': '+919800000002', 'VehicleModel': 'Nova S', 'VehicleVariant': 'Petrol MT', 'VehicleApplication': 'Hatchback', 'VIN': 'MA1NOVAS00000002', 'KilometerReading': 1875, 'ECUPartNumber': 'ECU-NV-2202', 'ECUName': 'Body Control Module', 'Severity': 'High', 'IssueType': 'Field Issue', 'CustodianNTID': 'CUS002', 'CustodianName': 'Engineer Kai', 'Description': 'Issue reproduced on test drive over cobblestone route.', 'Status': 'Investigation Started', 'CreatedBy': 'MGR001', 'CreatedAt': '2026-08-25T09:10:00+00:00', 'UpdatedBy': 'CUS002', 'UpdatedAt': '2026-08-26T05:50:00+00:00'},
        {'IncidentID': 'EFQ-20260820-0001', 'IncidentTitle': 'HVAC blower speed fluctuation after rain exposure', 'Date': '2026-08-20', 'OEM': 'Atlas Motors', 'CustomerComplaint': 'Blower speed changes automatically after vehicle wash or rain.', 'DealerName': 'Eastern Drive Care', 'DealerLocation': 'Bengaluru', 'DealerContact': '+919800000003', 'VehicleModel': 'Aster X', 'VehicleVariant': 'Petrol CVT', 'VehicleApplication': 'SUV', 'VIN': 'MA1ASTERX00000003', 'KilometerReading': 9240, 'ECUPartNumber': 'ECU-CL-3301', 'ECUName': 'Climate Control Unit', 'Severity': 'Medium', 'IssueType': 'Field Issue', 'CustodianNTID': 'CUS001', 'CustodianName': 'Engineer Ada', 'Description': 'Water ingress near connector cavity suspected by dealer.', 'Status': 'Closed', 'CreatedBy': 'OTH001', 'CreatedAt': '2026-08-20T07:00:00+00:00', 'UpdatedBy': 'CUS001', 'UpdatedAt': '2026-08-23T17:20:00+00:00'},
    ]


def _sample_resolutions() -> list[dict[str, Any]]:
    return [
        {'ResolutionID': 'RES-000001', 'IncidentID': 'EFQ-20260824-0001', 'InvestigationDetails': 'Reviewed cranking logs, connector drag test, and harness routing near battery tray.', 'RootCause': 'Terminal drag below lower specification on starter relay connector.', 'Recommendation': 'Replace affected terminal and retest starter current draw.', 'ProposedSolution': 'Introduce revised terminal retention check at end-of-line inspection.', 'CorrectiveAction': 'Reworked connector terminal and validated starter engagement over 20 cycles.', 'PreventiveAction': 'Added retention poke-yoke and updated supplier inspection checklist.', 'ValidationMethod': 'Cold soak crank validation', 'ValidationResult': 'Conditional Pass', 'ValidationDate': '2026-08-25', 'ResolutionOwner': 'CUS001', 'TargetDate': '2026-08-27', 'ResolutionDate': '', 'ResolutionStatus': 'Validation', 'Remarks': 'Awaiting additional environmental validation.', 'CreatedAt': '2026-08-24T12:00:00+00:00', 'UpdatedAt': '2026-08-25T11:15:00+00:00'},
        {'ResolutionID': 'RES-000002', 'IncidentID': 'EFQ-20260825-0001', 'InvestigationDetails': 'Confirmed cluster reset by vibration test and connector continuity scan.', 'RootCause': '', 'Recommendation': '', 'ProposedSolution': '', 'CorrectiveAction': '', 'PreventiveAction': '', 'ValidationMethod': '', 'ValidationResult': 'Not Tested', 'ValidationDate': '', 'ResolutionOwner': 'CUS002', 'TargetDate': '2026-08-29', 'ResolutionDate': '', 'ResolutionStatus': 'Investigation', 'Remarks': 'Supplier lot traceability under review.', 'CreatedAt': '2026-08-25T13:10:00+00:00', 'UpdatedAt': '2026-08-26T05:50:00+00:00'},
        {'ResolutionID': 'RES-000003', 'IncidentID': 'EFQ-20260820-0001', 'InvestigationDetails': 'Water spray test identified intermittent short on blower resistor connector seal.', 'RootCause': 'Connector seal lip folded during assembly causing moisture ingress.', 'Recommendation': 'Replace seal with revised part and retrain assembly operator.', 'ProposedSolution': 'Use revised connector seal with assembly guide sleeve.', 'CorrectiveAction': 'Replaced connector seal and updated plant work instruction.', 'PreventiveAction': 'Added audit checkpoint for seal seating during assembly.', 'ValidationMethod': 'Rain simulation and blower endurance test', 'ValidationResult': 'Pass', 'ValidationDate': '2026-08-23', 'ResolutionOwner': 'CUS001', 'TargetDate': '2026-08-23', 'ResolutionDate': '2026-08-23T17:20:00+00:00', 'ResolutionStatus': 'Closed', 'Remarks': 'Closed after validation sign-off.', 'CreatedAt': '2026-08-20T10:00:00+00:00', 'UpdatedAt': '2026-08-23T17:20:00+00:00'},
    ]


def _sample_activities() -> list[dict[str, Any]]:
    return [
        {'ActivityID': 'ACT-000001', 'IncidentID': 'EFQ-20260824-0001', 'Action': 'Incident Created', 'UserNTID': 'ADM001', 'UserName': 'Avery Stone', 'Timestamp': '2026-08-24T08:30:00+00:00'},
        {'ActivityID': 'ACT-000002', 'IncidentID': 'EFQ-20260824-0001', 'Action': 'Investigation Updated', 'UserNTID': 'CUS001', 'UserName': 'Engineer Ada', 'Timestamp': '2026-08-24T12:00:00+00:00'},
        {'ActivityID': 'ACT-000003', 'IncidentID': 'EFQ-20260824-0001', 'Action': 'Validation Updated', 'UserNTID': 'CUS001', 'UserName': 'Engineer Ada', 'Timestamp': '2026-08-25T11:15:00+00:00'},
        {'ActivityID': 'ACT-000004', 'IncidentID': 'EFQ-20260825-0001', 'Action': 'Incident Created', 'UserNTID': 'MGR001', 'UserName': 'Jordan Vale', 'Timestamp': '2026-08-25T09:10:00+00:00'},
        {'ActivityID': 'ACT-000005', 'IncidentID': 'EFQ-20260825-0001', 'Action': 'Investigation Updated', 'UserNTID': 'CUS002', 'UserName': 'Engineer Kai', 'Timestamp': '2026-08-26T05:50:00+00:00'},
        {'ActivityID': 'ACT-000006', 'IncidentID': 'EFQ-20260820-0001', 'Action': 'Incident Created', 'UserNTID': 'OTH001', 'UserName': 'Taylor Reed', 'Timestamp': '2026-08-20T07:00:00+00:00'},
        {'ActivityID': 'ACT-000007', 'IncidentID': 'EFQ-20260820-0001', 'Action': 'Root Cause Added', 'UserNTID': 'CUS001', 'UserName': 'Engineer Ada', 'Timestamp': '2026-08-21T13:00:00+00:00'},
        {'ActivityID': 'ACT-000008', 'IncidentID': 'EFQ-20260820-0001', 'Action': 'Incident Closed', 'UserNTID': 'CUS001', 'UserName': 'Engineer Ada', 'Timestamp': '2026-08-23T17:20:00+00:00'},
    ]


def _sample_lookup_rows() -> dict[str, list[dict[str, Any]]]:
    return {
        'OEMRegions': [
            {'OEM': 'Atlas Motors', 'Region': 'India West', 'IsActive': True},
            {'OEM': 'Nova Mobility', 'Region': 'India South', 'IsActive': True},
            {'OEM': 'Summit EV', 'Region': 'India North', 'IsActive': True},
        ],
        'ComplaintSuggestions': [
            {'SuggestionID': 'CS-001', 'ComplaintDescription': 'No crank on first ignition', 'Category': 'Starting', 'IsActive': True},
            {'SuggestionID': 'CS-002', 'ComplaintDescription': 'Cluster resets on rough road', 'Category': 'Electrical', 'IsActive': True},
            {'SuggestionID': 'CS-003', 'ComplaintDescription': 'HVAC blower speed fluctuation', 'Category': 'HVAC', 'IsActive': True},
        ],
        'ECUs': [
            {'ECUID': 'ECU-001', 'ECUName': 'Powertrain Control Module', 'ECUPartNumber': 'ECU-AX-1001', 'IsActive': True},
            {'ECUID': 'ECU-002', 'ECUName': 'Body Control Module', 'ECUPartNumber': 'ECU-NV-2202', 'IsActive': True},
            {'ECUID': 'ECU-003', 'ECUName': 'Climate Control Unit', 'ECUPartNumber': 'ECU-CL-3301', 'IsActive': True},
        ],
        'DetectionPhases': [
            {'PhaseID': 'DP-001', 'PhaseName': 'Dealer Entry', 'IsActive': True},
            {'PhaseID': 'DP-002', 'PhaseName': 'Field Return Analysis', 'IsActive': True},
            {'PhaseID': 'DP-003', 'PhaseName': 'Validation Bench', 'IsActive': True},
        ],
    }


class WorkbookManager:
    def __init__(self, file_path: Path) -> None:
        self.file_path = Path(file_path)
        self._lock = RLock()

    def ensure_workbook(self) -> None:
        with self._lock:
            self.file_path.parent.mkdir(parents=True, exist_ok=True)
            if not self.file_path.exists():
                workbook = self._create_workbook_with_seed_data()
                self._atomic_save(workbook)
                return
            try:
                workbook = load_workbook(self.file_path)
            except Exception as exc:
                raise RuntimeError(f'Workbook is corrupted or unreadable: {self.file_path}') from exc

            updated = False
            for sheet_name, columns in SHEET_SCHEMAS.items():
                if sheet_name not in workbook.sheetnames:
                    worksheet = workbook.create_sheet(sheet_name)
                    worksheet.append(columns)
                    updated = True
                    continue
                worksheet = workbook[sheet_name]
                if self._headers(worksheet) != columns:
                    existing_rows = self._rows_from_sheet(worksheet)
                    worksheet.delete_rows(1, worksheet.max_row)
                    worksheet.append(columns)
                    for row in existing_rows:
                        worksheet.append([row.get(column, '') for column in columns])
                    updated = True
            if updated:
                self._atomic_save(workbook)

    def read_rows(self, sheet_name: str) -> list[dict[str, Any]]:
        with self._lock:
            self.ensure_workbook()
            workbook = load_workbook(self.file_path)
            return self._rows_from_sheet(workbook[sheet_name])

    def replace_rows(self, sheet_name: str, rows: list[dict[str, Any]]) -> None:
        with self._lock:
            self.ensure_workbook()
            workbook = load_workbook(self.file_path)
            worksheet = workbook[sheet_name]
            columns = SHEET_SCHEMAS[sheet_name]
            worksheet.delete_rows(1, worksheet.max_row)
            worksheet.append(columns)
            for row in rows:
                worksheet.append([row.get(column, '') for column in columns])
            self._atomic_save(workbook)

    def append_row(self, sheet_name: str, row: dict[str, Any]) -> None:
        rows = self.read_rows(sheet_name)
        rows.append(row)
        self.replace_rows(sheet_name, rows)

    def _create_workbook_with_seed_data(self) -> Workbook:
        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)
        for sheet_name, columns in SHEET_SCHEMAS.items():
            worksheet = workbook.create_sheet(sheet_name)
            worksheet.append(columns)
        for row in _sample_users():
            workbook['Users'].append([row.get(column, '') for column in SHEET_SCHEMAS['Users']])
        for row in _sample_incidents():
            workbook['Incidents'].append([row.get(column, '') for column in SHEET_SCHEMAS['Incidents']])
        for row in _sample_resolutions():
            workbook['Resolutions'].append([row.get(column, '') for column in SHEET_SCHEMAS['Resolutions']])
        for row in _sample_activities():
            workbook['Activities'].append([row.get(column, '') for column in SHEET_SCHEMAS['Activities']])
        for sheet_name, rows in _sample_lookup_rows().items():
            for row in rows:
                workbook[sheet_name].append([row.get(column, '') for column in SHEET_SCHEMAS[sheet_name]])
        return workbook

    def _atomic_save(self, workbook: Workbook) -> None:
        backup_path = self.file_path.with_suffix('.bak.xlsx')
        if self.file_path.exists():
            copy2(self.file_path, backup_path)
        with NamedTemporaryFile(delete=False, suffix='.xlsx', dir=self.file_path.parent) as temp_file:
            temp_path = Path(temp_file.name)
        workbook.save(temp_path)
        temp_path.replace(self.file_path)

    @staticmethod
    def _headers(worksheet: Worksheet) -> list[str]:
        return [cell.value for cell in worksheet[1]] if worksheet.max_row else []

    @staticmethod
    def _rows_from_sheet(worksheet: Worksheet) -> list[dict[str, Any]]:
        if worksheet.max_row < 2:
            return []
        headers = [cell.value for cell in worksheet[1]]
        rows: list[dict[str, Any]] = []
        for values in worksheet.iter_rows(min_row=2, values_only=True):
            record = {header: WorkbookManager._serialize_cell(value) for header, value in zip(headers, values)}
            if any(value not in ('', None) for value in record.values()):
                rows.append(record)
        return rows

    @staticmethod
    def _serialize_cell(value: Any) -> Any:
        if isinstance(value, datetime):
            return value.astimezone(timezone.utc).replace(microsecond=0).isoformat()
        if isinstance(value, date):
            return value.isoformat()
        return value if value is not None else ''
